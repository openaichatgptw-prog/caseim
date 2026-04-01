# -*- coding: utf-8 -*-
"""
Clasifica registros de la hoja MEWPS_FORKLIFT en MARKETSHARE.xlsx:
- clasificacion: MEWPS | FORKLIFT | NA
- marca: extraída de texto o proveedor; NA si no aplica
- Estado: NUEVO | USADO | NA

Nota: en Excel los valores de texto "NA" se leen como NaN en pandas
      salvo que use read_excel(..., keep_default_na=False).
"""
from __future__ import annotations

import re
import unicodedata
import shutil
from pathlib import Path

import pandas as pd

SHEET = "MEWPS_FORKLIFT"
HEADER_ROW = 2  # 0-based en pandas: fila 2 = encabezados de datos
EXCEL_NAME = "MARKETSHARE.xlsx"


def _norm(s: str) -> str:
    if not isinstance(s, str):
        s = "" if s is None or (isinstance(s, float) and pd.isna(s)) else str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.upper()


def _score_mewp(t: str) -> int:
    score = 0
    patterns = [
        (r"PLATAFORMA\s+DE\s+TIJERA", 8),
        (r"PLATAFORMA\s+ELEVADORA", 6),
        (r"TIJERA\s+ELECTR", 7),
        (r"\bTIJERA\b", 3),
        (r"BRAZO\s+ARTICULADO", 8),
        (r"BOOM\s+LIFT", 6),
        (r"ARTICULATING\s+BOOM", 6),
        (r"SCISSOR\s+LIFT", 8),
        (r"ELECTRIC\s+SCISSOR", 8),
        (r"LIFTING\s+PLATFORM", 4),
        (r"ASCENSOR\s+TIPO\s+TIJERA", 8),
        (r"ALTURA\s+DE\s+TRABAJO", 5),
        (r"PLATAFORMA\s+DE\s+TRABAJO", 6),
        (r"AERIAL\s+WORK\s+PLATFORM", 8),
        (r"AWP\b", 3),
        (r"\bMEWP\b", 8),
        (r"MAN\s+LIFT", 4),
        (r"VERTICAL\s+LIFT", 4),
        (r"PERSONNEL\s+LIFT", 4),
        (r"ELEVADOR(ES)?\s+PLATAFORMA\s+DE\s+TIJERA", 8),
    ]
    for pat, w in patterns:
        if re.search(pat, t, re.I):
            score += w
    # marcas típicas MEWP (no forzar solo por marca; suma leve)
    brands_mewp = (
        r"\bGENIE\b",
        r"\bJLG\b",
        r"HAULOTTE",
        r"SKYJACK",
        r"SNORKEL",
        r"\bMANTALL\b",
        r"ALO\s+LIFT",
        r"ZOOMLION.*LIFT",  # ambiguo; bajo
    )
    for pat in brands_mewp:
        if re.search(pat, t, re.I):
            score += 2
    return score


def _score_forklift(t: str) -> int:
    score = 0
    patterns = [
        (r"CARRETILLA(S)?\s+ELEVADORA", 10),
        (r"MONTACARGA(S)?", 9),
        (r"\bFORKLIFT\b", 9),
        (r"MÁSTIL|MASTIL", 7),
        (r"TR[IÍ]PLEX", 6),
        (r"D[UÚ]PLEX", 5),
        (r"CONTRAPESAD", 6),
        (r"CONTRABALANCEAD", 6),
        (r"APILADORA", 5),
        (r"REACH\s+TRUCK", 7),
        (r"PALLET\s+TRUCK", 4),
        (r"TRANSPALET", 4),
        (r"NOBLELIFT", 5),
        (r"JUNGHEINRICH", 4),
        (r"\bLINDE\b", 4),
        (r"HYSTER", 4),
        (r"\bCROWN\b", 3),
        (r"HANGCHA", 4),
        (r"TOYOTA\s+8F", 3),
        (r"8FD|8FBN|8FG", 2),
    ]
    for pat, w in patterns:
        if re.search(pat, t, re.I):
            score += w
    return score


def clasificar(text: str) -> str:
    t = _norm(text)
    if not t.strip():
        return "NA"

    sm = _score_mewp(t)
    sf = _score_forklift(t)

    # Reglas de desempate fuertes
    if re.search(r"PLATAFORMA\s+DE\s+TIJERA|SCISSOR\s+LIFT|TIJERA\s+ELECTR", t, re.I) and not re.search(
        r"MONTACARGA|CARRETILLA\s+ELEVADORA", t, re.I
    ):
        sm += 5
    if re.search(r"MONTACARGA|CARRETILLA\s+ELEVADORA|FORKLIFT\s+TRUCK", t, re.I):
        sf += 4
    # "CARRETILLAS ELEVADORAS, LIFT PLATFORM" -> a menudo stacker/Mewp confuso: empujar por LIFT PLATFORM + sin MASTIL
    if "LIFT PLATFORM" in t and "MÁSTIL" not in t and "MASTIL" not in t and sf < 8:
        sm += 3

    if sm == 0 and sf == 0:
        return "NA"
    if sm > sf:
        return "MEWPS"
    if sf > sm:
        return "FORKLIFT"
    return "NA"


def _clean_marca(m: str) -> str | None:
    if not m:
        return None
    m = m.strip()
    m = re.sub(r"\s+", " ", m)
    upper = m.upper()
    bad = {
        "NO TIENE",
        "NO TIE NE",
        "NO TIENE.",
        "N/A",
        "NA",
        "SIN MARCA",
        "DESCONOCIDO",
        "NO APLICA",
        "SEGUN FACTURA",
        "NO",
    }
    if upper in bad or len(m) < 2:
        return None
    if re.match(r"^NO\s+TIENE", upper):
        return None
    # demasiado largo: probablemente captura incorrecta
    if len(m) > 48:
        m = m[:48].rsplit(" ", 1)[0]
    return m.upper()


def extraer_marca_descripcion(text: str) -> str | None:
    if not isinstance(text, str) or not text.strip():
        return None
    t = text.replace("\n", " ")
    patterns = [
        r"MARCA\s*[/]?\s*MODELO\s*:\s*([^,;\n]{2,40})",
        r"MARCA\s*:\s*([^,;\n]{2,40})",
        r"MARCA\s+C\s*:\s*([^,;\n]{2,40})",
        r"NOMBRE\s+COMERCIAL\s*:\s*[^,;]{0,60}MARCA\s*:\s*([^,;\n]{2,40})",
    ]
    for pat in patterns:
        m = re.search(pat, t, re.I)
        if m:
            cand = _clean_marca(m.group(1).strip())
            if cand:
                return cand
    return None


# Proveedor -> marca cuando el fabricante va en el nombre de la empresa
PROVEEDOR_MARCA = [
    (re.compile(r"TOYOTA\s+MATERIAL", re.I), "TOYOTA"),
    (re.compile(r"HYSTER\s*[-/]?\s*YALE|HYSTER", re.I), "HYSTER"),
    (re.compile(r"ZOOMLION", re.I), "ZOOMLION"),
    (re.compile(r"MANITOU", re.I), "MANITOU"),
    (re.compile(r"TEREX", re.I), "TEREX"),
    (re.compile(r"GENIE", re.I), "GENIE"),
    (re.compile(r"\bJLG\b|JLG\s+INDUSTRIES", re.I), "JLG"),
    (re.compile(r"HAULOTTE", re.I), "HAULOTTE"),
    (re.compile(r"CROWN\s+EQUIPMENT", re.I), "CROWN"),
    (re.compile(r"LINDE\s+MATERIAL", re.I), "LINDE"),
    (re.compile(r"KION\b", re.I), "KION"),
    (re.compile(r"NOBLELIFT", re.I), "NOBLELIFT"),
    (re.compile(r"HANGCHA", re.I), "HANGCHA"),
    (re.compile(r"MAXIMAL", re.I), "MAXIMAL"),
    (re.compile(r"EP\s+EQUIPMENT|EP\s+FORKLIFT", re.I), "EP"),
    (re.compile(r"ALO\s+GROUP|ALO-GROUP", re.I), "ALO LIFT"),
    (re.compile(r"CANNY\s+ELEVATOR", re.I), "CANNY"),
]


def marca_desde_proveedor(proveedor: str) -> str | None:
    if not isinstance(proveedor, str) or not str(proveedor).strip():
        return None
    p = str(proveedor).strip()
    for rx, marca in PROVEEDOR_MARCA:
        if rx.search(p):
            return marca
    return None


def inferir_estado(text: str) -> str:
    if not isinstance(text, str):
        return "NA"
    t = _norm(text)
    nuevo = bool(
        re.search(r"MERCANCIA\s+NUEVA|MERCANC[IÍ]A\s+NUEVA|ART[IÍ]CULO\s+NUEVO|NUEVA\s+DE\s+PRIMERA|NUEVA\s+Y\s+DE\s+PRIMERA", t)
    )
    usado = bool(
        re.search(
            r"MERCANCIA\s+USADA|MERCANC[IÍ]A\s+USADA|ART[IÍ]CULO\s+USADO|\bUSADA\b|\bUSADO\b|DE\s+SEGUNDA|"
            r"SEMINUEVO|REMANUFACTURADO|USO\s+O\s+DESTINO.*USAD",
            t,
        )
    )
    if nuevo and usado:
        return "NA"
    if nuevo:
        return "NUEVO"
    if usado:
        return "USADO"
    return "NA"


def main() -> None:
    root = Path(__file__).resolve().parent
    path = root / EXCEL_NAME
    if not path.exists():
        raise SystemExit(f"No se encontró {path}")

    df = pd.read_excel(path, sheet_name=SHEET, header=HEADER_ROW)
    desc_col = None
    for c in df.columns:
        if "mercanc" in str(c).lower():
            desc_col = c
            break
    if desc_col is None:
        raise SystemExit("No se encontró columna de descripción de mercancía")

    prov_col = "Proveedor" if "Proveedor" in df.columns else None

    desc_series = df[desc_col].fillna("").astype(str)
    prov_series = df[prov_col].fillna("").astype(str) if prov_col else pd.Series([""] * len(df))

    clasificaciones = []
    marcas = []
    estados = []

    for desc, prov in zip(desc_series, prov_series):
        cls = clasificar(desc)
        mar = extraer_marca_descripcion(desc)
        if mar is None:
            mar = marca_desde_proveedor(prov)
        if mar is None:
            mar = "NA"
        est = inferir_estado(desc)

        clasificaciones.append(cls)
        marcas.append(mar)
        estados.append(est)

    # Evitar duplicar columnas si se re-ejecuta el script
    for col in ("clasificacion", "marca", "Estado"):
        if col in df.columns:
            df = df.drop(columns=[col])

    df["clasificacion"] = clasificaciones
    df["marca"] = marcas
    df["Estado"] = estados

    from openpyxl import load_workbook

    wb = load_workbook(path)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Hoja {SHEET} no existe en el libro")
    ws = wb[SHEET]
    excel_header_row = HEADER_ROW + 1  # fila 3 en Excel (encabezados de columnas)

    start_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=excel_header_row, column=c).value
        if v is not None and str(v).strip().lower() == "clasificacion":
            start_col = c
            break

    if start_col is None:
        last_used = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=excel_header_row, column=c).value
            if v is not None and str(v).strip() != "":
                last_used = c
        start_col = last_used + 1
        ws.cell(row=excel_header_row, column=start_col, value="clasificacion")
        ws.cell(row=excel_header_row, column=start_col + 1, value="marca")
        ws.cell(row=excel_header_row, column=start_col + 2, value="Estado")

    first_data_row = excel_header_row + 1
    n = len(df)
    for i in range(n):
        r = first_data_row + i
        ws.cell(row=r, column=start_col, value=clasificaciones[i])
        ws.cell(row=r, column=start_col + 1, value=marcas[i])
        ws.cell(row=r, column=start_col + 2, value=estados[i])

    out_tmp = path.with_suffix(".tmp.xlsx")
    wb.save(out_tmp)
    wb.close()
    shutil.move(str(out_tmp), str(path))
    print(f"Guardado: {path} hoja {SHEET}")
    print("Resumen clasificacion:", pd.Series(clasificaciones).value_counts().to_string())
    print("Resumen Estado:", pd.Series(estados).value_counts().head(10).to_string())
    print("Marcas NA:", sum(1 for m in marcas if m == "NA"), "/", len(marcas))


if __name__ == "__main__":
    main()
